from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from faker import  Faker
import json

"""
可以读写Excel文件
"""

class DOExcel():

    # 构造器
    def __init__(self,filepath,sheetname):
        self.filepath = filepath
        self.sheetname = sheetname
        self.fake = Faker('zh_CN')

    def read_excel(self):
        wb = load_workbook(self.filepath)
        ws:Worksheet = wb[self.sheetname]
        data = []
        for row in ws.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data

    def get_random_data(self,my_data):
        """
        随机取其中一个数据
        TODO 数据首字符大写转换小写，去掉 .
        :param my_data:
        :return:
        """
        return self.fake.sentence(nb_words=1, ext_word_list=my_data)


    def write_excel(self):
        wb = Workbook()
        ws = wb.create_sheet(title=self.sheetname,index=0)

        ws.append(["case_id","method","url","params","expect_val","result_val","result"])

        mothod = "get"
        url = "http://39.107.96.138:3000/api/v1/topics"

        my_tab = ["ask", "share", "job", "good"]

        for x in range(2):
            case_id = x+1
            param_data = {
                "page": 1,
                "tab": self.get_random_data(my_tab),
                "limit": 1,
                "mdrander": "false"
            }
            # dict to str
            # param_data = str(param_data)
            param_data = json.dumps(param_data)
            ws.append([case_id,mothod,url,param_data])
        wb.save(self.filepath)





if __name__ == '__main__':
    doExcel = DOExcel('../data/topicdata.xlsx','indextopic')
    doExcel.write_excel()
    # all_data = doExcel.read_excel()
    # print(all_data)
